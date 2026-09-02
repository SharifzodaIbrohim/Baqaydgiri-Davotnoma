"""Export filtered results to .xlsx (Student ID as text)."""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _full_name(r: Dict[str, Any]) -> str:
    parts = [r.get("last_name") or "", r.get("first_name") or "", r.get("patronymic") or ""]
    return " ".join(p for p in parts if p).strip()


def build_results_xlsx(rows: List[Dict[str, Any]]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Натиҷаҳо"

    headers = [
        "Ном",
        "Student ID",
        "Мактаб",
        "Синф",
        "Ҷинс",
        "Фан",
        "Олимпиада",
        "Хол",
        "Макс",
        "Фоиз",
        "Статус",
        "Санаи бақайд",
        "Санаи хол",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A3328")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for i, r in enumerate(rows, 2):
        values = [
            _full_name(r),
            str(r.get("student_id") or ""),
            r.get("school") or "",
            r.get("class_name") or "",
            r.get("gender") or "",
            r.get("subject") or "",
            r.get("olympiad_title") or "",
            r.get("score"),
            r.get("max_score"),
            r.get("percent"),
            r.get("status") or "",
            r.get("created_at") or "",
            r.get("scored_at") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin
            if col == 2:
                cell.number_format = "@"
                cell.value = str(val)
            cell.alignment = Alignment(vertical="center")

    widths = [28, 22, 28, 10, 10, 22, 24, 10, 10, 10, 12, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{max(1, len(rows) + 1)}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
