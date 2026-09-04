"""Import students from Excel (.xlsx)."""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HEADER_MAP = {
    "насаб": "last_name",
    "last_name": "last_name",
    "фамилия": "last_name",
    "ном": "first_name",
    "first_name": "first_name",
    "имя": "first_name",
    "номи падар": "patronymic",
    "patronymic": "patronymic",
    "отчество": "patronymic",
    "таваллуд": "birth_date",
    "рӯзи таваллуд": "birth_date",
    "birth_date": "birth_date",
    "суроға": "address",
    "address": "address",
    "мактаб": "school",
    "муассиса": "school",
    "муассисаи таълимӣ": "school",
    "school": "school",
    "синф": "class_name",
    "class": "class_name",
    "class_name": "class_name",
    "омӯзгор": "teacher",
    "teacher": "teacher",
    "ҷинс": "gender",
    "gender": "gender",
    "фан": "subject",
    "фанни имтиҳонсупорӣ": "subject",
    "subject": "subject",
    "олимпиада": "olympiad_title",
    "намуди олимпиада": "olympiad_title",
    "унвони олимпиада": "olympiad_title",
    "olympiad_title": "olympiad_title",
    "сана": "olympiad_date",
    "санаи олимпиада": "olympiad_date",
    "olympiad_date": "olympiad_date",
}


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace("  ", " ")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def build_import_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Хонандагон"
    headers = [
        "Насаб", "Ном", "Номи падар", "Таваллуд", "Суроға", "Мактаб",
        "Синф", "Омӯзгор", "Ҷинс", "Фан", "Олимпиада", "Сана",
    ]
    fill = PatternFill("solid", fgColor="0A3328")
    font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="CFE0D6"),
        right=Side(style="thin", color="CFE0D6"),
        top=Side(style="thin", color="CFE0D6"),
        bottom=Side(style="thin", color="CFE0D6"),
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = 16
    ws.append([
        "Аҳмадов", "Али", "Қосимович", "01.01.2010", "ш. Бохтар",
        "Мактаби №1", "9а", "Каримова", "Мард", "МАТЕМАТИКА",
        "Олимпиадаи шаҳрӣ", "02.09.2026",
    ])
    note = wb.create_sheet("Дастур")
    note["A1"] = "Ҷинс: Мард ё Зан"
    note["A2"] = "Фан: бояд бо рӯйхати барнома мувофиқ бошад"
    note["A3"] = "Насаб, Ном, Мактаб, Синф, Ҷинс, Фан — ҳатмӣ"
    note["A4"] = "Сатри намунавӣ (сатри 2)-ро нест карда, маълумоти худро нависед"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_students_xlsx(file_bytes: bytes) -> Tuple[List[Dict[str, str]], List[str]]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Файл холӣ аст"]

    col_map: Dict[int, str] = {}
    for i, h in enumerate(header_row):
        key = HEADER_MAP.get(_norm_header(h))
        if key:
            col_map[i] = key

    required_present = {"last_name", "first_name", "school", "class_name", "gender", "subject"}
    found = set(col_map.values())
    missing_cols = required_present - found
    if missing_cols:
        return [], [
            "Сутунҳои ҳатмӣ намерасанд: " + ", ".join(sorted(missing_cols))
            + ". Намуна: Насаб, Ном, Мактаб, Синф, Ҷинс, Фан"
        ]

    out: List[Dict[str, str]] = []
    errors: List[str] = []
    for line_no, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        data = {k: "" for k in (
            "last_name", "first_name", "patronymic", "birth_date", "address",
            "school", "class_name", "teacher", "gender", "subject",
            "olympiad_title", "olympiad_date",
        )}
        for i, key in col_map.items():
            if i < len(row):
                data[key] = _cell_str(row[i])
        g = data["gender"]
        gl = g.lower()
        if gl in ("мард", "m", "male", "м"):
            data["gender"] = "Мард"
        elif gl in ("зан", "f", "female", "ж", "з"):
            data["gender"] = "Зан"
        miss = [k for k in ("last_name", "first_name", "school", "class_name", "gender", "subject") if not data.get(k)]
        if miss:
            errors.append(f"Сатри {line_no}: майдонҳои холӣ — {', '.join(miss)}")
            continue
        if data["gender"] not in ("Мард", "Зан"):
            errors.append(f"Сатри {line_no}: Ҷинс бояд Мард ё Зан бошад (ҳоло: {data['gender']})")
            continue
        out.append(data)
    return out, errors
